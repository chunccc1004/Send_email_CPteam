# %%
# 메일 보낼 상대의 이메일 주소 리스트(you_list)를 만들기 위한 모듈
import os
import re
import openpyxl
from openpyxl import load_workbook

# 메일을 보내기 위한 모듈
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

user_id = input('아이디를 입력해주세요(네이버 아이디만 가능)')
user_password = input('비밀번호를 입력해주세요')
user_email = user_id+'@naver.com'
error_email_list = []

smtp_info = dict({"smtp_server" : "smtp.naver.com", # SMTP 서버 주소
                  "smtp_user_id" : user_id, 
                  "smtp_user_pw" : user_password,
                  "smtp_port" : 587}) # SMTP 서버 포트
                  
def send_email(smtp_info, msg):
    with smtplib.SMTP(smtp_info["smtp_server"], smtp_info["smtp_port"]) as server:
        # TLS 보안 연결
        server.starttls() 
        # 로그인
        server.login(smtp_info["smtp_user_id"], smtp_info["smtp_user_pw"])

        # 로그인 된 서버에 이메일 전송
        response = server.sendmail(msg['from'], msg['to'], msg.as_string()) # 메시지를 보낼때는 .as_string() 메소드를 사용해서 문자열로 바꿔줍니다.

        # 이메일을 성공적으로 보내면 결과는 {}
        if not response:
            print('이메일을 성공적으로 보냈습니다.')
        else:
            print(response)
  
               
# me == my email address
# you == recipient's email address
me = user_email
you_list = []

ecxel_adress = os.path.abspath('total_email_adress.xlsx') #파일의 주소 불러오는 것
load_wb = load_workbook(ecxel_adress)
load_ws = load_wb['email_adress']
order = 1
while 1 :
    cell_val = str(load_ws.cell(order,1).value)
    var2 = cell_val.find('@')
    var = cell_val.find('@')
    order = order + 1
    if var == -1 :
        break
    you_list.append(cell_val)
you_list.append(user_email)

# Create message container - the correct MIME type is multipart/alternative.
subject = input('메일 제목을 입력해주세요')
for j in you_list :
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = me
    msg['To'] = j

    # Create the body of the message (a plain-text and an HTML version).
    html_start = """\
    <html>
      <head></head>
      <body>
      """

    html_final = """
      </body>
    </html>
    """

    filename = "index.txt"
    text=""
    with open (filename, 'rt', encoding='utf8') as myfile:  # 파일 불러내기
        for myline in myfile:                 # 모든 파일 1줄씩 읽기
            text = text+myline                # 1줄씩 출력하기
    html_index = text
    html = html_start + html_index + html_final


    #개선해야할 점 : html은 메모장에서 추출해서 사용하되 img파일만 따로
    #사용 방법은 네이버에서 메일보내는걸로 만든 후에 html 코드 복사 붙여넣기 메모장으로 해서 놓고 이미지 파일 놓으면 exe로 실행시켜서 메일 보내기
    #보내야하는 이메일 주소도 크롤링 해야하고 메일 주소 이상한건 제외하는 것도 해야하고 중복 메일도 체크해야할 듯 -> 완료
    #Record the MIME types of both parts - text/plain and text/html.
    #내게 보내기로 확인 후 전체에게 보내는게 좋을 듯
    part2 = MIMEText(html, 'html')

    # Attach parts into message container.
    # According to RFC 2046, the last part of a multipart message, in this case
    # the HTML message, is best and preferred.
    msg.attach(part2)
    try:
        send_email(smtp_info, msg)
    except:
        error_email_list.append(j)
if len(error_email_list) > 0 :
    print("에러가 난 이메일 목록")
    print(error_email_list)