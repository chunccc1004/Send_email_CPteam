# %%
import os
import re
import openpyxl
from openpyxl import load_workbook

total_adress_list = [] # 찾는 목표
count2 = 0
# 엑셀 파일 찾기위해서 파일 이름들 보는 것 -> 완성
path_dir = os.getcwd()
file_list = os.listdir(path_dir)
ecxel_name_list = []
for i in file_list:
    var1 = i.find('.xlsx')
    if var1 < 0:
        continue
    ecxel_name_list.append(i)
# print(ecxel_name_list)

for i in ecxel_name_list:
    # 엑셀 파일에서 시트 이름 불러오기
    ecxel_adress = os.path.abspath(i) #파일의 주소 불러오는 것
    load_wb = load_workbook(ecxel_adress)
    sheet_name_list = load_wb.sheetnames
    # 찾은 sheet 이름을 이용해서 각 sheet 내의 이메일이 적히 열 찾기
    for j in sheet_name_list:
        # print(j)
        Row = 1
        Col = 1
        count = 0
        load_ws = load_wb[j]
        while 1:
            if Col < 20:
                Col = Col + 1
            else:
                Col = 1
                Row = Row + 1
            cell_val = str(load_ws.cell(Row, Col).value)
            var2 = cell_val.find('@')
            count = count + 1
            if var2 != -1 :
                break
            if count > 400:
                break
        # 찾은 열을 이용해서 각 주소 한 리스트에 넣기
        Row = 1
        count = 0
        old = count2
        while 1:
            cell_val = str(load_ws.cell(Row, Col).value)
            var3 = cell_val.find('@')
            if var3 != -1:
                total_adress_list.append(cell_val)
                count2 = count2 + 1
                count = 0
            elif var3 == -1:
                count = count+1
            Row = Row+1
            if count > 20:
                break
        #print(count2 - old) 

    # print(i) # 각 엑셀 명 확인
    # print(sheet_name_list) # 각 시트 명 확인
    # print(total_adress_list) # 총 저장된 이메일 주소 확인
    # print(len(total_adress_list)) # 총 이메일 주소 갯수 확인   
    
#중복 이메일 제거
total_adress_set = set(total_adress_list)
total_adress_list = list(total_adress_set)

# 셀이 여러개인 엑셀에서 주소값만 추출해서 리스트로 만드는 것 성공!

# 모은 이메일을 하나의 엑셀 파일에 정리
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'email_adress'

count3 = 1
for x in total_adress_list :
    sheet.cell(row = count3, column = 1).value = x
    count3 = count3 + 1 
wb.save('total_email_adress.xlsx')

print('완료')
# print(total_adress_list)
# print(len(total_adress_list))
out = input('종료하고 싶으면 아무런 키나 눌러주세요')
if out != None :
    print('종료')

# %%
